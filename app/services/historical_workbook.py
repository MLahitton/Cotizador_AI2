import hashlib
import zipfile
from enum import StrEnum
from pathlib import Path

from openpyxl import load_workbook
from pydantic import BaseModel, Field

from app.models.historical_quote import (
    HistoricalQuoteIssue,
    HistoricalQuoteIssueSeverity,
    HistoricalQuoteSource,
)

OOXML_CONTENT_TYPES = "[Content_Types].xml"
OOXML_WORKBOOK = "xl/workbook.xml"
OLE_CDFV2_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
ZIP_MAGIC = b"PK"
QUOTE_SHEET_NAME = "COTIZACIÓN"


class HistoricalWorkbookType(StrEnum):
    OOXML = "ooxml"
    OLE_CDFV2 = "ole_cdfv2"
    UNKNOWN = "unknown"


class HistoricalWorkbookIssueCode(StrEnum):
    UNSUPPORTED_CONTAINER = "HISTORICAL_UNSUPPORTED_CONTAINER"
    INVALID_OOXML = "HISTORICAL_INVALID_OOXML"
    MISSING_QUOTE_SHEET = "HISTORICAL_MISSING_QUOTE_SHEET"
    EMPTY_FILE = "HISTORICAL_EMPTY_FILE"
    UNKNOWN_FORMAT = "HISTORICAL_UNKNOWN_FORMAT"
    WORKBOOK_OPEN_ERROR = "HISTORICAL_WORKBOOK_OPEN_ERROR"


class HistoricalWorkbookInspection(BaseModel):
    source: HistoricalQuoteSource
    is_processable: bool = False
    sheet_names: list[str] = Field(default_factory=list)
    has_quote_sheet: bool = False
    issues: list[HistoricalQuoteIssue] = Field(default_factory=list)


def calculate_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def inspect_historical_workbook(
    path: Path,
    *,
    source_index: int | None = None,
    source_path: str | None = None,
) -> HistoricalWorkbookInspection:
    resolved_path = Path(path)
    sha256 = calculate_sha256(resolved_path)
    file_format = _file_format(resolved_path)
    size = resolved_path.stat().st_size
    source = HistoricalQuoteSource(
        file_name=resolved_path.name,
        sha256=sha256,
        file_format=file_format,
        workbook_type=HistoricalWorkbookType.UNKNOWN.value,
        source_path=source_path,
        source_index=source_index,
    )
    issues: list[HistoricalQuoteIssue] = []

    if size == 0:
        issues.append(
            _issue(
                HistoricalWorkbookIssueCode.EMPTY_FILE,
                "Historical workbook file is empty.",
                HistoricalQuoteIssueSeverity.ERROR,
            )
        )
        return HistoricalWorkbookInspection(source=source, issues=issues)

    header = _read_header(resolved_path)
    if header.startswith(OLE_CDFV2_MAGIC):
        source.workbook_type = HistoricalWorkbookType.OLE_CDFV2.value
        issues.append(
            _issue(
                HistoricalWorkbookIssueCode.UNSUPPORTED_CONTAINER,
                "OLE/CDFV2 historical workbooks are detected but not parsed yet.",
                HistoricalQuoteIssueSeverity.ERROR,
            )
        )
        return HistoricalWorkbookInspection(source=source, issues=issues)

    if header.startswith(ZIP_MAGIC):
        return _inspect_ooxml_candidate(resolved_path, source, issues)

    issues.append(
        _issue(
            HistoricalWorkbookIssueCode.UNKNOWN_FORMAT,
            "Historical workbook format could not be identified from content.",
            HistoricalQuoteIssueSeverity.ERROR,
        )
    )
    return HistoricalWorkbookInspection(source=source, issues=issues)


def _inspect_ooxml_candidate(
    path: Path,
    source: HistoricalQuoteSource,
    issues: list[HistoricalQuoteIssue],
) -> HistoricalWorkbookInspection:
    try:
        with zipfile.ZipFile(path) as archive:
            names = set(archive.namelist())
    except zipfile.BadZipFile:
        issues.append(
            _issue(
                HistoricalWorkbookIssueCode.INVALID_OOXML,
                "File starts as ZIP but is not a valid OOXML ZIP container.",
                HistoricalQuoteIssueSeverity.ERROR,
            )
        )
        return HistoricalWorkbookInspection(source=source, issues=issues)

    if OOXML_CONTENT_TYPES not in names or OOXML_WORKBOOK not in names:
        issues.append(
            _issue(
                HistoricalWorkbookIssueCode.INVALID_OOXML,
                "ZIP container does not contain the minimal XLSX workbook structure.",
                HistoricalQuoteIssueSeverity.ERROR,
            )
        )
        return HistoricalWorkbookInspection(source=source, issues=issues)

    source.workbook_type = HistoricalWorkbookType.OOXML.value
    try:
        workbook = load_workbook(
            path,
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except Exception:
        issues.append(
            _issue(
                HistoricalWorkbookIssueCode.WORKBOOK_OPEN_ERROR,
                "OOXML workbook could not be opened safely in read-only mode.",
                HistoricalQuoteIssueSeverity.ERROR,
            )
        )
        return HistoricalWorkbookInspection(source=source, issues=issues)

    try:
        sheet_names = list(workbook.sheetnames)
    finally:
        workbook.close()

    has_quote_sheet = QUOTE_SHEET_NAME in sheet_names
    if not has_quote_sheet:
        issues.append(
            _issue(
                HistoricalWorkbookIssueCode.MISSING_QUOTE_SHEET,
                "Workbook does not contain a COTIZACIÓN sheet.",
                HistoricalQuoteIssueSeverity.WARNING,
            )
        )

    return HistoricalWorkbookInspection(
        source=source,
        is_processable=True,
        sheet_names=sheet_names,
        has_quote_sheet=has_quote_sheet,
        issues=issues,
    )


def _read_header(path: Path, size: int = 16) -> bytes:
    with path.open("rb") as file:
        return file.read(size)


def _file_format(path: Path) -> str:
    suffix = path.suffix.lower().lstrip(".")
    return suffix or "unknown"


def _issue(
    code: HistoricalWorkbookIssueCode,
    message: str,
    severity: HistoricalQuoteIssueSeverity,
) -> HistoricalQuoteIssue:
    return HistoricalQuoteIssue(
        code=code.value,
        message=message,
        severity=severity,
    )
