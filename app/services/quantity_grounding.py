from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Protocol

from openpyxl import load_workbook

from app.models.common import ExtractionStatus
from app.models.gemini_enrichment import (
    GeminiElementEnrichment,
    GeminiEnrichmentResult,
)

SOURCE_INDEPENDENT_QUANTITY = "SOURCE_INDEPENDENT_QUANTITY"
SOURCE_GROUNDED_QUANTITY = SOURCE_INDEPENDENT_QUANTITY
MODEL_EVIDENCE_QUANTITY = "MODEL_EVIDENCE_QUANTITY"
MODEL_EXPLICIT_QUANTITY = "MODEL_EXPLICIT_QUANTITY"
MODEL_INFERRED_QUANTITY = "MODEL_INFERRED_QUANTITY"
SOURCE_INDEPENDENT_GROUNDED_WINS = "SOURCE_INDEPENDENT_GROUNDED_WINS"
SOURCE_GROUNDED_WINS = SOURCE_INDEPENDENT_GROUNDED_WINS
SOURCE_MODEL_AGREE = "SOURCE_MODEL_AGREE"
MODEL_QUANTITY_NOT_SOURCE_GROUNDED = "MODEL_QUANTITY_NOT_SOURCE_GROUNDED"
NO_SOURCE_GROUNDING_AVAILABLE = "NO_SOURCE_GROUNDING_AVAILABLE"
QUANTITY_GROUNDING_CONFLICT = "QUANTITY_GROUNDING_CONFLICT"
NO_MODEL_QUANTITY = "NO_MODEL_QUANTITY"
SPREADSHEET_CELL = "SPREADSHEET_CELL"

_NUMBER = r"\d+(?:[.,]\d+)?"
_QUANTITY_RE = re.compile(
    rf"\b(?:cantidad(?:\s+total)?|cant\.?|cnt|qty|unidades|und)\s*:?\s*({_NUMBER})\b",
    flags=re.IGNORECASE,
)
_FORMAL_REFERENCE_RE = re.compile(r"\b([A-Z]{1,4})[\s._-]*(\d{1,4})([A-Z]?)\b")


class QuantitySourceFileSpec(Protocol):
    path: Path
    mime_type: str


@dataclass(frozen=True)
class QuantityGroundingCandidate:
    temporary_id: str
    reference: str | None
    value: int | float
    source_id: str | None
    source_file_name: str | None
    page_number: int | None
    sheet_name: str | None
    cell_range: str | None
    region: object | None
    raw_text: str
    field_path: str | None
    origin: str
    source_type: str
    status: ExtractionStatus


@dataclass(frozen=True)
class QuantityGroundingDecision:
    temporary_id: str
    reference: str | None
    original_quantity: str | int | float | None
    final_quantity: str | int | float | None
    action: str
    reason: str
    source_candidates: tuple[QuantityGroundingCandidate, ...]
    model_evidence_candidates: tuple[QuantityGroundingCandidate, ...]
    selected_source_candidate: QuantityGroundingCandidate | None = None


def build_source_independent_quantity_candidates(
    file_specs: list[QuantitySourceFileSpec],
) -> dict[str, list[QuantityGroundingCandidate]]:
    candidates: dict[str, list[QuantityGroundingCandidate]] = {}
    for index, spec in enumerate(file_specs, start=1):
        source_id = f"source-{index}"
        if spec.mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            _extend_candidates(candidates, _spreadsheet_quantity_candidates(spec.path, source_id))
    return candidates


def validate_enrichment_quantities(
    enrichment: GeminiEnrichmentResult,
    source_candidates_by_reference: dict[str, list[QuantityGroundingCandidate]] | None = None,
) -> tuple[GeminiEnrichmentResult, list[QuantityGroundingDecision]]:
    elements: list[GeminiElementEnrichment] = []
    decisions: list[QuantityGroundingDecision] = []
    warnings = list(enrichment.warnings)
    source_candidates_by_reference = source_candidates_by_reference or {}

    for element in enrichment.elements:
        updated, decision = validate_element_quantity(
            element,
            source_candidates_by_reference=source_candidates_by_reference,
        )
        elements.append(updated)
        decisions.append(decision)
        if decision.action in {"REPLACE", "MARK_AMBIGUOUS", "DOWNGRADE"}:
            warnings.append(
                f"{decision.reason}: {element.temporary_id} "
                f"reference={element.reference!r} quantity={element.quantity!r}."
            )

    return (
        GeminiEnrichmentResult(elements=elements, warnings=warnings, usage=enrichment.usage),
        decisions,
    )


def validate_element_quantity(
    element: GeminiElementEnrichment,
    *,
    source_candidates_by_reference: dict[str, list[QuantityGroundingCandidate]] | None = None,
) -> tuple[GeminiElementEnrichment, QuantityGroundingDecision]:
    source_candidates = tuple(
        _source_candidates_for_element(element, source_candidates_by_reference or {})
    )
    model_evidence_candidates = tuple(_model_evidence_quantity_candidates(element))
    model_quantity = _numeric_value(element.quantity)

    if model_quantity is None:
        return element, _decision(
            element,
            final_quantity=element.quantity,
            action="KEEP",
            reason=NO_MODEL_QUANTITY,
            source_candidates=source_candidates,
            model_evidence_candidates=model_evidence_candidates,
        )

    distinct_source_values = _distinct_values(source_candidates)
    if len(distinct_source_values) > 1:
        updated = _replace_quantity(
            element,
            quantity=element.quantity,
            status=ExtractionStatus.AMBIGUOUS,
            note=(
                f"{QUANTITY_GROUNDING_CONFLICT}: model quantity {element.quantity!r} "
                "does not match a single independent source quantity candidate."
            ),
        )
        return updated, _decision(
            element,
            final_quantity=element.quantity,
            action="MARK_AMBIGUOUS",
            reason=QUANTITY_GROUNDING_CONFLICT,
            source_candidates=source_candidates,
            model_evidence_candidates=model_evidence_candidates,
        )

    matching_candidates = [
        candidate
        for candidate in source_candidates
        if _same_number(candidate.value, model_quantity)
    ]
    if matching_candidates:
        return element, _decision(
            element,
            final_quantity=element.quantity,
            action="KEEP",
            reason=SOURCE_MODEL_AGREE,
            source_candidates=source_candidates,
            model_evidence_candidates=model_evidence_candidates,
            selected_source_candidate=matching_candidates[0],
        )

    if len(distinct_source_values) == 1:
        grounded_quantity = distinct_source_values[0]
        updated = _replace_quantity(
            element,
            quantity=grounded_quantity,
            status=ExtractionStatus.EXPLICIT,
            note=(
                f"{SOURCE_INDEPENDENT_GROUNDED_WINS}: replaced model quantity "
                f"{element.quantity!r} with independent source quantity "
                f"{grounded_quantity!r}."
            ),
        )
        return updated, _decision(
            element,
            final_quantity=grounded_quantity,
            action="REPLACE",
            reason=SOURCE_INDEPENDENT_GROUNDED_WINS,
            source_candidates=source_candidates,
            model_evidence_candidates=model_evidence_candidates,
            selected_source_candidate=source_candidates[0],
        )

    updated = _replace_quantity(
        element,
        quantity=element.quantity,
        status=ExtractionStatus.INFERRED,
        confidence=_downgraded_confidence(element.confidence),
        note=(
            f"{NO_SOURCE_GROUNDING_AVAILABLE}: quantity {element.quantity!r} "
            "has no independent source representation with a quantity label-value pair."
        ),
    )
    return updated, _decision(
        element,
        final_quantity=element.quantity,
        action="DOWNGRADE",
        reason=NO_SOURCE_GROUNDING_AVAILABLE,
        source_candidates=source_candidates,
        model_evidence_candidates=model_evidence_candidates,
    )


def _model_evidence_quantity_candidates(
    element: GeminiElementEnrichment,
) -> list[QuantityGroundingCandidate]:
    candidates: list[QuantityGroundingCandidate] = []
    for index, evidence in enumerate(element.evidence, start=1):
        text = evidence.text or evidence.visual_description
        if not text:
            continue
        for match in _QUANTITY_RE.finditer(text):
            candidates.append(
                QuantityGroundingCandidate(
                    temporary_id=element.temporary_id,
                    reference=element.reference,
                    value=_parse_number(match.group(1)),
                    source_id=evidence.source_id,
                    source_file_name=None,
                    page_number=evidence.page_number,
                    sheet_name=evidence.sheet_name,
                    cell_range=evidence.cell_range,
                    region=evidence.region,
                    raw_text=match.group(0),
                    field_path=f"evidence[{index}]",
                    origin=MODEL_EVIDENCE_QUANTITY,
                    source_type=MODEL_EVIDENCE_QUANTITY,
                    status=ExtractionStatus.EXPLICIT,
                )
            )
    return candidates


def _source_candidates_for_element(
    element: GeminiElementEnrichment,
    source_candidates_by_reference: dict[str, list[QuantityGroundingCandidate]],
) -> list[QuantityGroundingCandidate]:
    reference = _canonical_reference(element.reference)
    if reference is None:
        return []
    return list(source_candidates_by_reference.get(reference, []))


def _replace_quantity(
    element: GeminiElementEnrichment,
    *,
    quantity: str | int | float | None,
    status: ExtractionStatus,
    note: str,
    confidence: float | None = None,
) -> GeminiElementEnrichment:
    missing = list(element.missing_or_unknown)
    if note.split(":", 1)[0] not in missing:
        missing.append(note.split(":", 1)[0])
    return element.model_copy(
        update={
            "quantity": quantity,
            "quantity_status": status,
            "quantity_confidence": element.confidence if confidence is None else confidence,
            "quantity_notes": _append_note(element.quantity_notes, note),
            "missing_or_unknown": missing,
            "notes": _append_note(element.notes, note),
        }
    )


def _decision(
    element: GeminiElementEnrichment,
    *,
    final_quantity: str | int | float | None,
    action: str,
    reason: str,
    source_candidates: tuple[QuantityGroundingCandidate, ...],
    model_evidence_candidates: tuple[QuantityGroundingCandidate, ...],
    selected_source_candidate: QuantityGroundingCandidate | None = None,
) -> QuantityGroundingDecision:
    return QuantityGroundingDecision(
        temporary_id=element.temporary_id,
        reference=element.reference,
        original_quantity=element.quantity,
        final_quantity=final_quantity,
        action=action,
        reason=reason,
        source_candidates=source_candidates,
        model_evidence_candidates=model_evidence_candidates,
        selected_source_candidate=selected_source_candidate,
    )


def _spreadsheet_quantity_candidates(
    path: Path,
    source_id: str,
) -> list[QuantityGroundingCandidate]:
    candidates: list[QuantityGroundingCandidate] = []
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            quantity_columns: set[int] = set()
            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                cells = ["" if value is None else str(value).strip() for value in row]
                for cell_index, cell in enumerate(cells, start=1):
                    if _is_quantity_label(cell):
                        quantity_columns.add(cell_index)
                references = _references_in_cells(cells)
                if not references:
                    continue
                for quantity_value, cell_index, raw_text in _row_quantity_values(
                    cells,
                    quantity_columns,
                ):
                    for reference in references:
                        candidates.append(
                            QuantityGroundingCandidate(
                                temporary_id="",
                                reference=reference,
                                value=quantity_value,
                                source_id=source_id,
                                source_file_name=path.name,
                                page_number=None,
                                sheet_name=sheet.title,
                                cell_range=f"{sheet.cell(row_index, cell_index).coordinate}",
                                region=None,
                                raw_text=raw_text,
                                field_path=None,
                                origin=SPREADSHEET_CELL,
                                source_type=SOURCE_INDEPENDENT_QUANTITY,
                                status=ExtractionStatus.EXPLICIT,
                            )
                        )
    finally:
        workbook.close()
    return candidates


def _references_in_cells(cells: list[str]) -> list[str]:
    references: list[str] = []
    for cell in cells:
        for match in _FORMAL_REFERENCE_RE.finditer(cell.upper()):
            reference = _canonical_reference(match.group(0))
            if reference and reference not in references:
                references.append(reference)
    return references


def _row_quantity_values(
    cells: list[str],
    quantity_columns: set[int],
) -> list[tuple[int | float, int, str]]:
    values: list[tuple[int | float, int, str]] = []
    for index, cell in enumerate(cells, start=1):
        for match in _QUANTITY_RE.finditer(cell):
            values.append((_parse_number(match.group(1)), index, match.group(0)))
        if _is_quantity_label(cell) and index < len(cells):
            next_value = _numeric_value(cells[index])
            if next_value is not None:
                values.append((next_value, index + 1, f"{cell} {cells[index]}"))
        if index in quantity_columns:
            value = _numeric_value(cell)
            if value is not None:
                values.append((value, index, f"CANTIDAD {cell}"))
    return values


def _is_quantity_label(value: str) -> bool:
    normalized = value.casefold().replace(".", "").strip()
    return normalized in {
        "cantidad",
        "cantidad total",
        "cant",
        "cnt",
        "qty",
        "unidades",
        "und",
    }


def _extend_candidates(
    candidates: dict[str, list[QuantityGroundingCandidate]],
    items: list[QuantityGroundingCandidate],
) -> None:
    for item in items:
        reference = _canonical_reference(item.reference)
        if reference is None:
            continue
        candidates.setdefault(reference, []).append(item)


def _numeric_value(value: str | int | float | None) -> int | float | None:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int | float):
        return value
    match = re.fullmatch(_NUMBER, value.strip())
    if match is None:
        return None
    return _parse_number(value)


def _canonical_reference(reference: str | None) -> str | None:
    if not reference:
        return None
    value = reference.strip().upper()
    match = _FORMAL_REFERENCE_RE.fullmatch(value)
    if match is None:
        return None
    prefix, number, suffix = match.groups()
    return f"{prefix}-{int(number):02d}{suffix}"


def _parse_number(value: str) -> int | float:
    parsed = float(value.replace(",", "."))
    return int(parsed) if parsed.is_integer() else parsed


def _same_number(left: int | float, right: int | float) -> bool:
    return abs(float(left) - float(right)) < 0.000001


def _distinct_values(
    candidates: tuple[QuantityGroundingCandidate, ...],
) -> list[int | float]:
    values: list[int | float] = []
    for candidate in candidates:
        if not any(_same_number(candidate.value, value) for value in values):
            values.append(candidate.value)
    return values


def _downgraded_confidence(confidence: float | None) -> float | None:
    if confidence is None:
        return None
    return min(confidence, 0.5)


def _append_note(current: str | None, note: str) -> str:
    if not current:
        return note
    if note in current:
        return current
    return f"{current}\n{note}"
